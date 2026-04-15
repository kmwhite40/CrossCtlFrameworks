"""Slim xlsx → SQLite importer used by Concord Reader.

Avoids the full ETL pipeline's provenance tables (workbook_versions,
control_history, mapping_history) which are Postgres-specific in practice.
Writes directly into the ATTACHed `ccf` database.
"""

from __future__ import annotations

import asyncio
import json
from pathlib import Path

import openpyxl
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncEngine, create_async_engine

from ..etl.frameworks import CORE_HEADERS, FRAMEWORKS, classify_header

ASSESSMENT_SHEET = "SP.800-53Ar5_assessment"


async def ingest_into_sqlite(engine: AsyncEngine, xlsx_path: Path) -> dict[str, int]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    stats = {"controls": 0, "mappings": 0, "worksheets": 0}
    try:
        async with engine.begin() as conn:
            # Seed frameworks catalog
            for spec in FRAMEWORKS:
                await conn.execute(
                    text(
                        """
                        INSERT OR IGNORE INTO ccf.frameworks(
                            code,
                            name,
                            family,
                            description
                        )
                        VALUES (:c, :n, :f, :d)
                        """
                    ),
                    {"c": spec.code, "n": spec.name, "f": spec.family, "d": spec.description},
                )
            await conn.execute(
                text(
                    """
                    INSERT OR IGNORE INTO ccf.frameworks(
                        code,
                        name,
                        family,
                        description
                    )
                    VALUES ('OTHER', 'Other / Misc', 'Other', 'Unclassified columns')
                    """
                )
            )
            fw_rows = (await conn.execute(text("SELECT id, code FROM ccf.frameworks"))).all()
            fw_ids = {r.code: r.id for r in fw_rows}

            # Clear existing catalog rows
            await conn.execute(text("DELETE FROM ccf.framework_mappings"))
            await conn.execute(text("DELETE FROM ccf.controls"))
            await conn.execute(text("DELETE FROM ccf.worksheet_rows"))
            await conn.execute(text("DELETE FROM ccf.worksheets"))

            # Assessment sheet
            if ASSESSMENT_SHEET in wb.sheetnames:
                ws = wb[ASSESSMENT_SHEET]
                headers: list[str] = []
                seen: set[str] = set()
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        headers = [str(h or f"col_{j}") for j, h in enumerate(row)]
                        continue
                    if not any(c is not None and str(c).strip() for c in row):
                        continue
                    record = dict(zip(headers, row, strict=False))
                    ident = record.get("identifier") or ""
                    ident = str(ident).strip() if ident else ""
                    if not ident:
                        continue
                    if ident in seen:
                        ident = f"{ident}#row{i}"
                    seen.add(ident)

                    payload = {k: v for k, v in record.items() if v is not None and str(v).strip()}
                    r = await conn.execute(
                        text("""
                        INSERT INTO ccf.controls(identifier, sort_as, control_name,
                            description, assessment_objective, examine, interview, test,
                            fisma_low, fisma_mod, fisma_high, audit_payload, source_row)
                        VALUES (:ident, :sort, :name, :desc, :obj, :ex, :iv, :tst,
                                :l, :m, :h, :payload, :row)
                    """),
                        {
                            "ident": ident,
                            "sort": record.get("sort-as"),
                            "name": record.get("control-name"),
                            "desc": record.get("Security Control Description"),
                            "obj": record.get("assessment-objective"),
                            "ex": record.get("EXAMINE"),
                            "iv": record.get("INTERVIEW"),
                            "tst": record.get("TEST"),
                            "l": 1
                            if str(record.get("FISMA Low") or "").strip().lower()
                            in {"x", "yes", "y", "1"}
                            else 0,
                            "m": 1
                            if str(record.get("FISMA Mod") or "").strip().lower()
                            in {"x", "yes", "y", "1"}
                            else 0,
                            "h": 1
                            if str(record.get("FISMA High") or "").strip().lower()
                            in {"x", "yes", "y", "1"}
                            else 0,
                            "payload": json.dumps(payload, default=str),
                            "row": i,
                        },
                    )
                    ctl_id = r.lastrowid
                    stats["controls"] += 1

                    for header, value in record.items():
                        if header in CORE_HEADERS or value is None:
                            continue
                        v = str(value).strip()
                        if not v:
                            continue
                        await conn.execute(
                            text(
                                """
                                INSERT INTO ccf.framework_mappings(
                                    control_id,
                                    framework_id,
                                    column_key,
                                    value
                                )
                                VALUES (:cid, :fid, :col, :val)
                                """
                            ),
                            {
                                "cid": ctl_id,
                                "fid": fw_ids.get(classify_header(header) or "OTHER"),
                                "col": header,
                                "val": v,
                            },
                        )
                        stats["mappings"] += 1

            stats["worksheets"] = len(wb.sheetnames) - (
                1 if ASSESSMENT_SHEET in wb.sheetnames else 0
            )
    finally:
        wb.close()
    return stats


def run_ingest(dsn: str, xlsx_path: Path) -> dict[str, int]:
    """Synchronous wrapper for the launcher."""
    engine = create_async_engine(dsn)
    try:
        return asyncio.run(ingest_into_sqlite(engine, xlsx_path))
    finally:
        asyncio.run(engine.dispose())
