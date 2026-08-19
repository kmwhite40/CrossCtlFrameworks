"""Workbook ingestion pipeline.

Flow per run:
  1. SHA-256 the source; upsert ccf_audit.workbook_versions (content-addressed).
  2. Open a ccf.ingestion_runs row linked to the version.
  3. Validate assessment-tab headers against the packaged header contract.
  4. Ingest assessment -> ccf.controls + ccf.framework_mappings.
     Rows without identifier are quarantined in ccf_audit.rejected_rows.
  5. Snapshot the authoritative row set into ccf_audit.control_history and
     ccf_audit.mapping_history (one snapshot per workbook version).
  6. Ingest every other sheet into ccf.worksheets / ccf.worksheet_rows.
  7. Refresh the controls.search_vector column.
  8. Close the run with per-sheet stats.
"""

from __future__ import annotations

import hashlib
import re
from collections.abc import Iterable
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import openpyxl
from slugify import slugify
from sqlalchemy import delete, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from ..logging import get_logger
from ..models import (
    POAM,
    Control,
    ControlFamily,
    ControlHistory,
    ControlImplementation,
    Framework,
    FrameworkMapping,
    IngestionRun,
    MappingHistory,
    RejectedRow,
    Risk,
    WorkbookVersion,
    Worksheet,
    WorksheetRow,
)
from .frameworks import CORE_HEADERS, FRAMEWORKS, classify_header
from .validate import HeaderContractError, load_contract, validate_headers

log = get_logger(__name__)

ASSESSMENT_SHEET = "SP.800-53Ar5_assessment"
BOOL_STRINGS_TRUE = {"x", "yes", "y", "true", "t", "1"}
BOOL_STRINGS_FALSE = {"no", "n", "false", "f", "0"}
# Family label may carry the code as a leading OR trailing parenthetical:
# "(AC) Access Control" or "Access Control (AC)". Search (not match) so either
# position is recognised, and prefer the text outside the parenthetical as the name.
FAMILY_RE = re.compile(r"\(([A-Z]{2,3})\)")


def _clean(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return v


def _coerce_bool(v: Any) -> bool | None:
    cleaned = _clean(v)
    if cleaned is None:
        return None
    s = str(cleaned).strip().lower()
    if s in BOOL_STRINGS_TRUE:
        return True
    if s in BOOL_STRINGS_FALSE:
        return False
    return None


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


async def _upsert_workbook_version(
    session: AsyncSession, xlsx_path: Path, sha: str
) -> WorkbookVersion:
    existing = (
        await session.execute(select(WorkbookVersion).where(WorkbookVersion.sha256 == sha))
    ).scalar_one_or_none()
    if existing:
        return existing
    wv = WorkbookVersion(
        sha256=sha,
        source_path=str(xlsx_path),
        size_bytes=xlsx_path.stat().st_size,
    )
    session.add(wv)
    await session.flush()
    return wv


async def _seed_frameworks(session: AsyncSession) -> dict[str, int]:
    existing = {f.code: f.id for f in (await session.execute(select(Framework))).scalars().all()}
    for spec in FRAMEWORKS:
        if spec.code in existing:
            continue
        session.add(
            Framework(
                code=spec.code,
                name=spec.name,
                family=spec.family,
                description=spec.description,
            )
        )
    if "OTHER" not in existing:
        session.add(
            Framework(
                code="OTHER",
                name="Other / Misc",
                family="Other",
                description="Unclassified columns",
            )
        )
    await session.flush()
    return {f.code: f.id for f in (await session.execute(select(Framework))).scalars().all()}


async def _ensure_family(
    session: AsyncSession, raw: str | None, cache: dict[str, int]
) -> int | None:
    if not raw:
        return None
    label = raw.strip()
    m = FAMILY_RE.search(label)
    if m:
        code = m.group(1)
        # The name is whatever text sits outside the "(XX)" parenthetical.
        name = FAMILY_RE.sub("", label).strip(" -").strip().title() or label
    else:
        code, name = label[:16].upper(), label
    if code in cache:
        return cache[code]
    obj = (
        await session.execute(select(ControlFamily).where(ControlFamily.code == code))
    ).scalar_one_or_none()
    if obj is None:
        obj = ControlFamily(code=code, name=name)
        session.add(obj)
        await session.flush()
    cache[code] = obj.id
    return obj.id


def _sheet_headers(ws: Any) -> list[str]:
    """The first (header) row of a worksheet, normalized to strings."""
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        return [str(h) if h is not None else f"col_{idx}" for idx, h in enumerate(row)]
    return []


def _iter_sheet_rows(ws: Any) -> Iterable[tuple[int, list[str], tuple[Any, ...]]]:
    headers: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h) if h is not None else f"col_{idx}" for idx, h in enumerate(row)]
            continue
        if not any(c is not None and str(c).strip() for c in row):
            continue
        yield i, headers, row


async def _ingest_assessment(
    session: AsyncSession,
    ws: Any,
    framework_ids: dict[str, int],
    run: IngestionRun,
    workbook_version: WorkbookVersion | None,
) -> dict[str, int]:
    stats = {
        "rows": 0,
        "mappings": 0,
        "rejected": 0,
        "controls_created": 0,
        "controls_updated": 0,
        "controls_removed": 0,
        "controls_retained": 0,
    }
    family_cache: dict[str, int] = {}

    # Validate the header contract BEFORE touching existing data — this runs even
    # when the sheet is header-only (no data rows), and a contract failure aborts
    # the run without wiping the catalog.
    header_set = set(_sheet_headers(ws))
    diff = validate_headers(header_set, load_contract())
    if diff.added:
        log.info("ingest.header_drift", new_headers_count=len(diff.added))

    # Mappings are derived data with no inbound foreign keys, so replacing them
    # wholesale is safe and keeps the load simple.
    await session.execute(delete(FrameworkMapping))

    # Controls are NOT replaceable. ``controls.id`` is a sequence, so deleting
    # and reloading mints new ids, and four tables point at the old ones:
    # framework_mappings (CASCADE), control_implementations (RESTRICT),
    # poams (SET NULL) and risks (SET NULL). The RESTRICT makes the delete fail
    # outright once any SSP work exists; the two SET NULLs would succeed and
    # silently sever every POA&M and risk from its control.
    #
    # So match on the natural key instead — ``controls.identifier`` is UNIQUE —
    # and update in place. An id, once handed out, is never reused or reissued.
    existing_controls = {
        c.identifier: c for c in (await session.execute(select(Control))).scalars()
    }
    workbook_identifiers: set[str] = set()

    core_map = {
        "Sequence Control": "sequence_control",
        "sort-as": "sort_as",
        "Rev 5 Assurance Control?": "assurance_control",
        "NIST SP 800-53R5  Control": "control_number",
        "AP Acronym (from IGAP Control Export on RMF KS)": "ap_acronym",
        "OPD?": "opd",
        "control-name": "control_name",
        "Security Control Description": "description",
        "Security Control Discussion": "discussion",
        "NIST SP 800-53 Rev. 5 related controls": "related_controls",
        "Owner": "owner",
        "Overall Control Type": "overall_control_type",
        "Implemented By": "implemented_by",
        "assessment-objective": "assessment_objective",
        "EXAMINE": "examine",
        "INTERVIEW": "interview",
        "TEST": "test",
        "FISMA Low": "fisma_low",
        "FISMA Mod": "fisma_mod",
        "FISMA High": "fisma_high",
    }
    bool_fields = {"opd", "fisma_low", "fisma_mod", "fisma_high"}

    mapping_batch: list[FrameworkMapping] = []
    history_mappings: list[MappingHistory] = []
    history_controls: list[ControlHistory] = []
    seen_identifiers: set[str] = set()

    for row_idx, headers, row in _iter_sheet_rows(ws):

        record = dict(zip(headers, row, strict=False))
        identifier = _clean(record.get("identifier"))
        if not identifier:
            session.add(
                RejectedRow(
                    run_id=run.id,
                    sheet=ASSESSMENT_SHEET,
                    row_index=row_idx,
                    rule="missing_identifier",
                    payload={k: str(v) for k, v in record.items() if v is not None},
                )
            )
            stats["rejected"] += 1
            continue
        identifier = str(identifier)
        if identifier in seen_identifiers:
            identifier = f"{identifier}#row{row_idx}"
        seen_identifiers.add(identifier)

        family_id = await _ensure_family(
            session,
            _clean(record.get("family")),
            family_cache,
        )

        audit_payload = {k: v for k, v in record.items() if _clean(v) is not None}
        workbook_identifiers.add(identifier)
        ctl = existing_controls.get(identifier)
        if ctl is None:
            ctl = Control(identifier=identifier)
            session.add(ctl)
            stats["controls_created"] += 1
        else:
            stats["controls_updated"] += 1
        ctl.family_id = family_id
        ctl.source_row = row_idx
        ctl.audit_payload = audit_payload
        for header, attr in core_map.items():
            val = record.get(header)
            if attr in bool_fields:
                setattr(ctl, attr, _coerce_bool(val))
            else:
                cleaned = _clean(val)
                setattr(ctl, attr, str(cleaned) if cleaned is not None else None)
        await session.flush()
        stats["rows"] += 1

        if workbook_version is not None:
            history_controls.append(
                ControlHistory(
                    identifier=identifier,
                    workbook_version_id=workbook_version.id,
                    payload=audit_payload,
                )
            )

        for header, value in record.items():
            if header in CORE_HEADERS:
                continue
            v = _clean(value)
            if v is None:
                continue
            fw_code = classify_header(header) or "OTHER"
            mapping_batch.append(
                FrameworkMapping(
                    control_id=ctl.id,
                    framework_id=framework_ids.get(fw_code),
                    column_key=header,
                    value=str(v),
                )
            )
            if workbook_version is not None:
                history_mappings.append(
                    MappingHistory(
                        identifier=identifier,
                        workbook_version_id=workbook_version.id,
                        column_key=header,
                        value=str(v),
                    )
                )
            stats["mappings"] += 1

        if len(mapping_batch) >= 500:
            session.add_all(mapping_batch)
            mapping_batch.clear()
            await session.flush()
        if len(history_mappings) >= 1000:
            session.add_all(history_mappings)
            history_mappings.clear()
            await session.flush()

    if mapping_batch:
        session.add_all(mapping_batch)
    if history_controls:
        session.add_all(history_controls)
    if history_mappings:
        session.add_all(history_mappings)
    await session.flush()

    await _retire_absent_controls(session, existing_controls, workbook_identifiers, stats)

    # Postgres-only: refresh tsvector. SQLite (Reader) skips this.
    dialect = session.bind.dialect.name if session.bind else ""
    if dialect == "postgresql":
        await session.execute(
            text("""
            UPDATE ccf.controls SET search_vector =
              setweight(to_tsvector('english', coalesce(identifier,'')), 'A') ||
              setweight(to_tsvector('english', coalesce(control_name,'')), 'A') ||
              setweight(to_tsvector('english', coalesce(assessment_objective,'')), 'B') ||
              setweight(to_tsvector('english', coalesce(description,'')), 'C') ||
              setweight(to_tsvector('english', coalesce(discussion,'')), 'D')
        """)
        )
    return stats



#: Tables that hold a reference to ``ccf.controls`` which a delete would not
#: clean up. ``framework_mappings`` is absent deliberately: it cascades, and the
#: load has already replaced it. Each entry is (model, attribute).
_CONTROL_DEPENDANTS: tuple[tuple[type[Any], str], ...] = (
    (ControlImplementation, "control_id"),
    (POAM, "control_id"),
    (Risk, "control_id"),
)


async def _retire_absent_controls(
    session: AsyncSession,
    existing: dict[str, Control],
    seen: set[str],
    stats: dict[str, int],
) -> None:
    """Remove controls the workbook no longer contains — but only safely.

    A control that disappears from the catalog should not linger. A control that
    someone has written an implementation, POA&M or risk against must not vanish
    underneath them: ``poams`` and ``risks`` are ON DELETE SET NULL, so deleting
    one would quietly cut the link rather than refuse. Those are kept and counted,
    and the run says so.

    Doing nothing at all was not an option either — that is how a renamed control
    ends up in the catalog twice.
    """
    absent = [c for identifier, c in existing.items() if identifier not in seen]
    if not absent:
        return

    absent_ids = [c.id for c in absent]
    referenced: set[int] = set()
    for model, attr in _CONTROL_DEPENDANTS:
        column = getattr(model, attr)
        rows = (
            await session.execute(select(column).where(column.in_(absent_ids)).distinct())
        ).all()
        referenced.update(cid for (cid,) in rows if cid is not None)

    removable = [c for c in absent if c.id not in referenced]
    for control in removable:
        await session.delete(control)
    stats["controls_removed"] = len(removable)
    stats["controls_retained"] = len(absent) - len(removable)

    if stats["controls_retained"]:
        log.warning(
            "ingest.controls_retained",
            count=stats["controls_retained"],
            reason="referenced by an implementation, POA&M or risk",
            identifiers=sorted(c.identifier for c in absent if c.id in referenced)[:20],
        )
    if removable:
        log.info("ingest.controls_removed", count=len(removable))
    await session.flush()


async def _ingest_generic_sheet(
    session: AsyncSession,
    sheet_name: str,
    ws: Any,
) -> dict[str, int]:
    stats = {"rows": 0}
    slug = slugify(sheet_name, max_length=240) or f"sheet-{abs(hash(sheet_name))}"

    existing = (
        await session.execute(select(Worksheet).where(Worksheet.name == sheet_name))
    ).scalar_one_or_none()
    if existing is not None:
        await session.delete(existing)
        await session.flush()

    headers: list[str] = []
    rows_out: list[WorksheetRow] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h) if h is not None else f"col_{idx}" for idx, h in enumerate(row)]
            continue
        if not any(c is not None and str(c).strip() for c in row):
            continue
        payload = {
            h: (v if isinstance(v, (int, float, bool)) else str(v).strip())
            for h, v in zip(headers, row, strict=False)
            if v is not None and str(v).strip()
        }
        if not payload:
            continue
        rows_out.append(WorksheetRow(row_index=i, payload=payload))
        stats["rows"] += 1

    sheet = Worksheet(
        name=sheet_name,
        slug=slug,
        headers=headers,
        row_count=stats["rows"],
        rows=rows_out,
    )
    session.add(sheet)
    await session.flush()
    return stats


async def ingest_workbook(session: AsyncSession, xlsx_path: Path) -> IngestionRun:
    log.info("ingest.start", path=str(xlsx_path))
    sha = _sha256(xlsx_path)
    dialect = session.bind.dialect.name if session.bind else ""

    workbook_version = None
    if dialect == "postgresql":
        workbook_version = await _upsert_workbook_version(session, xlsx_path, sha)

    run = IngestionRun(
        source_file=str(xlsx_path),
        sha256=sha,
        workbook_version_id=(workbook_version.id if workbook_version else None),
    )
    session.add(run)
    await session.flush()

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        framework_ids = await _seed_frameworks(session)

        per_sheet: dict[str, dict[str, int]] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if sheet_name == ASSESSMENT_SHEET:
                per_sheet[sheet_name] = await _ingest_assessment(
                    session,
                    ws,
                    framework_ids,
                    run,
                    workbook_version,
                )
            else:
                per_sheet[sheet_name] = await _ingest_generic_sheet(session, sheet_name, ws)
            log.info("ingest.sheet", sheet=sheet_name, **per_sheet[sheet_name])

        run.finished_at = datetime.now(UTC)
        run.status = "succeeded"
        run.stats = {"sheets": per_sheet, "sha256": sha}
        await session.flush()

        try:
            from ..catalog.report import run_and_store  # noqa: PLC0415

            await run_and_store(session)
        except Exception as exc:  # advisory only — never fail ingest on reconciliation
            log.warning("catalog.reconcile_failed", error=str(exc))

        return run
    except HeaderContractError as e:
        run.finished_at = datetime.now(UTC)
        run.status = "failed"
        run.stats = {"error": "header_contract", "detail": str(e)}
        await session.flush()
        raise
    except Exception as e:
        run.finished_at = datetime.now(UTC)
        run.status = "failed"
        run.stats = {"error": "exception", "detail": str(e)[:500]}
        await session.flush()
        raise
    finally:
        wb.close()
