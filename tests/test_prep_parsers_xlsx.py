"""XLSX parser — per-sheet pages, column headers, coordinate fidelity."""

from __future__ import annotations

import io
import logging

import openpyxl

from ccf.prep.parsers import dispatch
from ccf.prep.parsers.xlsx import parse_xlsx


def _xlsx_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Account Review"
    ws.append(["Account", "Last Reviewed", "Reviewer"])
    ws.append(["svc-backup", "2026-07-01", "system owner"])
    ws2 = wb.create_sheet("Audit Retention")
    ws2.append(["Log Source", "Retention"])
    ws2.append(["firewall", "365 days"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_each_sheet_becomes_its_own_page_named_by_section_title() -> None:
    doc = parse_xlsx(_xlsx_bytes(), "inventory.xlsx")
    assert [p.section_title for p in doc.pages] == ["Account Review", "Audit Retention"]
    assert [p.page_number for p in doc.pages] == [1, 2]


def test_cell_carries_column_header_and_sheet_scoped_table_id() -> None:
    doc = parse_xlsx(_xlsx_bytes(), "inventory.xlsx")
    line = next(x for x in doc.iter_lines() if x.content == "2026-07-01")
    assert line.cell_label == "Last Reviewed"
    assert line.row_index == 1
    assert line.col_index == 1
    assert line.table_id == "Account Review"


def test_section_path_is_the_sheet_name() -> None:
    doc = parse_xlsx(_xlsx_bytes(), "inventory.xlsx")
    line = next(x for x in doc.iter_lines() if x.content == "365 days")
    assert line.section_path == "Audit Retention"


def test_blank_cells_are_skipped_but_do_not_shift_coordinates() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A", "B", "C"])
    ws.append(["one", None, "three"])
    buf = io.BytesIO()
    wb.save(buf)
    doc = parse_xlsx(buf.getvalue(), "s.xlsx")
    line = next(x for x in doc.iter_lines() if x.content == "three")
    assert line.col_index == 2
    assert line.cell_label == "C"


def test_numeric_cells_are_stringified() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Days"])
    ws.append([365])
    buf = io.BytesIO()
    wb.save(buf)
    doc = parse_xlsx(buf.getvalue(), "n.xlsx")
    assert any(x.content == "365" for x in doc.iter_lines())


def test_dispatch_routes_xlsx() -> None:
    doc = dispatch(_xlsx_bytes(), "inventory.xlsx", None)
    assert doc.parser_name == "xlsx"


def test_corrupt_workbook_returns_an_error_document_rather_than_raising() -> None:
    doc = parse_xlsx(b"not a real xlsx", "bad.xlsx")
    assert doc.success is False
    assert doc.error is not None


def test_corrupt_workbook_failure_path_survives_an_enabled_logger() -> None:
    """Regression for the reserved-LogRecord-attribute defect.

    tests/conftest.py runs Alembic migrations, and migrations/env.py calls
    logging.config.fileConfig(..., disable_existing_loggers=True), which disables
    this module's stdlib logger for the whole pytest session. A logger.warning()
    call on a *disabled* logger short-circuits before building a LogRecord, so a
    bug in the log call itself (e.g. passing `extra={"filename": ...}`, which
    collides with LogRecord's own reserved `filename` attribute and raises
    KeyError from stdlib logging) would never fire and this suite would pass
    while the same call raises outside pytest. Force the logger enabled so this
    test genuinely exercises the log call instead of being masked by that side
    effect.
    """
    logger = logging.getLogger("ccf.prep.parsers.xlsx")
    original = logger.disabled
    logger.disabled = False
    try:
        doc = parse_xlsx(b"not a real xlsx", "bad.xlsx")
    finally:
        logger.disabled = original
    assert doc.success is False
    assert doc.error is not None
